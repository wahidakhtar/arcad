from __future__ import annotations

import base64
import gzip
import io
import json
import logging
import subprocess
from datetime import date, datetime, timedelta, timezone
from urllib.parse import urlparse

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.config import get_settings

_log = logging.getLogger("arcad.backup")

_DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.file"]

# ─── Site schemas to export (non-recurring project schemas) ──────────────────

_SITE_SCHEMAS = [
    ("MD", "schema_md"),
    ("MA", "schema_ma"),
    ("MC", "schema_mc"),
    ("MI", "schema_mi"),
]

_BB_SCHEMA = ("BB", "schema_bb")


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _drive_service():
    s = get_settings()
    if s.google_service_account_json_b64:
        raw = base64.b64decode(s.google_service_account_json_b64).decode()
    elif s.google_service_account_json:
        raw = s.google_service_account_json
    else:
        raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON_B64 not configured")
    info = json.loads(raw)
    creds = service_account.Credentials.from_service_account_info(info, scopes=_DRIVE_SCOPES)
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _upload(service, name: str, data: bytes, mimetype: str, folder_id: str) -> str:
    meta = {"name": name, "parents": [folder_id]}
    media = MediaIoBaseUpload(io.BytesIO(data), mimetype=mimetype, resumable=False)
    f = service.files().create(body=meta, media_body=media, fields="id").execute()
    return f["id"]


def _cleanup_old(service, folder_id: str, days: int = 30) -> int:
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")
    query = f"'{folder_id}' in parents and createdTime < '{cutoff}' and trashed = false"
    result = service.files().list(q=query, fields="files(id, name)").execute()
    files = result.get("files", [])
    for f in files:
        service.files().delete(fileId=f["id"]).execute()
        _log.info("backup: deleted old file %s (%s)", f["name"], f["id"])
    return len(files)


# ─── pg_dump ──────────────────────────────────────────────────────────────────

def _pg_dump() -> bytes:
    db_url = get_settings().effective_db_url
    parsed = urlparse(db_url)
    env = {
        "PGPASSWORD": parsed.password or "",
        "PATH": "/usr/bin:/bin:/usr/local/bin",
    }
    cmd = [
        "pg_dump",
        "-h", parsed.hostname or "localhost",
        "-p", str(parsed.port or 5432),
        "-U", parsed.username or "postgres",
        "-d", (parsed.path or "").lstrip("/"),
        "--no-password",
        "--format=plain",
        "--encoding=UTF8",
    ]
    result = subprocess.run(cmd, capture_output=True, env=env, timeout=300)
    if result.returncode != 0:
        raise RuntimeError(f"pg_dump failed: {result.stderr.decode()[:500]}")
    return gzip.compress(result.stdout)


# ─── Excel ────────────────────────────────────────────────────────────────────

def _header_style(ws, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="8B1A1A")
    font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def _write_sheet(ws, db: Session, query: str) -> None:
    rows = db.execute(text(query))
    keys = list(rows.keys())
    _header_style(ws, keys)
    for row in rows:
        ws.append([str(v) if v is not None else "" for v in row])


def _build_excel(db: Session) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Non-recurring site projects ───────────────────────────────────────────
    for label, schema in _SITE_SCHEMAS:
        try:
            ws = wb.create_sheet(f"{label} Sites")
            _write_sheet(ws, db, f"SELECT * FROM {schema}.sites ORDER BY id")
        except Exception:
            _log.warning("backup: skipping sheet '%s Sites' (table may not exist)", label)
            wb.remove(ws)

    # ── BB sites (recurring) ──────────────────────────────────────────────────
    label, schema = _BB_SCHEMA
    try:
        ws = wb.create_sheet(f"{label} Sites")
        _write_sheet(ws, db, f"SELECT * FROM {schema}.sites ORDER BY id")
    except Exception:
        _log.warning("backup: skipping BB Sites sheet")
        wb.remove(ws)

    # ── Assignments ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Assignments")
    _write_sheet(ws, db, "SELECT * FROM schema_ops.subcon_assignments ORDER BY id")

    # ── Transactions ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Transactions")
    _write_sheet(ws, db, "SELECT * FROM schema_acc.transactions ORDER BY id")

    # ── POs ───────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("POs")
    _write_sheet(ws, db, "SELECT * FROM schema_acc.pos ORDER BY id")

    # ── Invoices ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Invoices")
    _write_sheet(ws, db, "SELECT * FROM schema_acc.invoices ORDER BY id")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Main entry point ─────────────────────────────────────────────────────────

def run_backup(db: Session) -> dict:
    s = get_settings()
    if not s.backup_drive_folder_id:
        raise RuntimeError("BACKUP_DRIVE_FOLDER_ID not configured")

    today = date.today().isoformat()
    service = _drive_service()
    result: dict = {}

    # pg_dump → Drive
    _log.info("backup: running pg_dump")
    dump_bytes = _pg_dump()
    dump_name = f"arcad_restore_{today}.sql.gz"
    fid = _upload(service, dump_name, dump_bytes, "application/gzip", s.backup_drive_folder_id)
    _log.info("backup: uploaded %s (%d bytes, id=%s)", dump_name, len(dump_bytes), fid)
    result["restore_file"] = dump_name
    result["restore_size_kb"] = round(len(dump_bytes) / 1024, 1)

    # Excel → Drive
    _log.info("backup: building Excel")
    xl_bytes = _build_excel(db)
    xl_name = f"arcad_excel_{today}.xlsx"
    fid = _upload(service, xl_name, xl_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", s.backup_drive_folder_id)
    _log.info("backup: uploaded %s (%d bytes, id=%s)", xl_name, len(xl_bytes), fid)
    result["excel_file"] = xl_name
    result["excel_size_kb"] = round(len(xl_bytes) / 1024, 1)

    # Cleanup files older than 30 days
    deleted = _cleanup_old(service, s.backup_drive_folder_id, days=30)
    result["deleted_old"] = deleted

    return result
